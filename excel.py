import os
import subprocess
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment

import tkinter as tk
from tkinter import messagebox, filedialog

import analysis

DBLOC = f"{os.environ['USERPROFILE']}\\.GradeX Analyzer\\.result-data\\"
os.makedirs(DBLOC, exist_ok=True)

# def refresh_sheet_list(sheet_listbox):
#     sheet_listbox.delete(0, tk.END)
#     result_files = os.listdir(DBLOC)
#     for file in result_files:
#         if file.endswith(".xlsx"):
#             sheet_listbox.insert(tk.END, file)

def refresh_sheet_list(sheet_listbox):
    sheet_listbox.delete(0, tk.END)
    result_files = [f for f in os.listdir(DBLOC) if f.endswith(".xlsx")]
    
    # Sort files by modification time (newest first)
    result_files.sort(key=lambda f: os.path.getmtime(os.path.join(DBLOC, f)), reverse=True)
    
    for file in result_files:
        sheet_listbox.insert(tk.END, file)

def save_to_excel(data, course, semester, batch, sheet_listbox):
    filename = f"{course}_Sem{semester}_{batch}.xlsx"
    filepath = os.path.join(DBLOC, filename)

    # Ensure output directory exists
    os.makedirs(DBLOC, exist_ok=True)

    # Convert data to DataFrames
    student_df = pd.DataFrame(data[1:], columns=data[0])
    analysis_df = analysis.perform_analysis(student_df)

    # Create a new Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    ws.merge_cells("A1:P1")  
    ws["A1"] = "ACROPOLIS INSTITUTE OF TECHNOLOGY AND RESEARCH, INDORE"
    ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:P2")
    ws["A2"] = "Faculty of Computer Application"
    ws["A2"].font = Font(name="Times New Roman", size=12, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:P3")
    ws["A3"] = f"ANALYSIS OF RESULT BATCH ({int(batch) + 2000}) {course} Sem - {semester}"
    ws["A3"].font = Font(name="Times New Roman", size=12, bold=True)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    # Add student_df (Student Data)
    for row in dataframe_to_rows(student_df, index=False, header=True):
        ws.append(row)

    last_row = ws.max_row # Find the last row after student data 

    # Add 5 empty rows
    for _ in range(5):
        ws.append([])

    # Merge and set "Result Analysis" heading dynamically
    ws.merge_cells(f"A{last_row + 6}:I{last_row + 6}")
    ws[f"A{last_row + 6}"] = f"Result Analysis {course} {semester} Sem"
    ws[f"A{last_row + 6}"].font = Font(name="Times New Roman", size=14, bold=True)
    ws[f"A{last_row + 6}"].alignment = Alignment(horizontal="center")

    # Add analysis_df (Analysis Data)
    for row in dataframe_to_rows(analysis_df, index=True, header=True):  
        ws.append(row)

    # Center-align all text in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None: 
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row == 4 or cell.row == (last_row + 7):  # Bold only specific rows
                    cell.font = Font(bold=True)


    # Handle PermissionError when saving file
    while True:
        try:
            wb.save(filepath)
            break  # If successful, break loop
        except PermissionError:
            messagebox.showerror("File Error", f"Close '{filename}' and try again.")
            import time
            time.sleep(3)  # Wait before retrying

    refresh_sheet_list(sheet_listbox) 

    
def download(sheet_listbox):
    selected = sheet_listbox.curselection()
    if not selected:
        messagebox.showerror("Error", "No sheet selected!")
        return
    filename = sheet_listbox.get(selected[0])
    src_path = os.path.join(DBLOC, filename)
    dest_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")],initialfile=filename)
    if dest_path:
        os.replace(src_path, dest_path)
        messagebox.showinfo("Success", "Sheet downloaded successfully!")
        refresh_sheet_list(sheet_listbox)

def delete(sheet_listbox):
    selected = sheet_listbox.curselection()
    if not selected:
        messagebox.showerror("Error", "No sheet selected!")
        return
    filename = sheet_listbox.get(selected[0])
    filepath = os.path.join(DBLOC, filename)
    if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete {filename}?"):
        os.remove(filepath)
        refresh_sheet_list(sheet_listbox)
        messagebox.showinfo("Deleted", "Sheet deleted successfully!")

def view(sheet_listbox):
    selected = sheet_listbox.curselection()
    if not selected:
        messagebox.showerror("Error", "No sheet selected!")
        return
    filename = sheet_listbox.get(selected[0])
    filepath = os.path.join(DBLOC, filename)
    subprocess.run(["start", "", filepath], shell=True)

def filter(event, sheet_listbox, filter_entry):
    """Filters the listbox based on user input (Course, Semester, Branch)."""
    query = filter_entry.get().strip().lower()
    
    if not query:  
        refresh_sheet_list(sheet_listbox)  # Show all files if query is empty
        return  

    sheet_listbox.delete(0, tk.END)
    result_files = os.listdir(DBLOC)

    for file in result_files:
        if file.endswith(".xlsx"):
            parts = file.replace(".xlsx", "").split("_")  
            if len(parts) >= 3:
                course_name, sem_part, branch_name = parts[0].lower(), parts[1].lower(), parts[2].lower()

                course_match = query == course_name  
                sem_match = query in sem_part
                branch_match = query in branch_name

                # Show file only if it matches Course, Semester, or Branch
                if course_match or sem_match or branch_match:
                    sheet_listbox.insert(tk.END, file)